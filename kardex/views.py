from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import is_aware, make_naive
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers

from .models import MovimientoInventario
from productos.models import Producto


# =======================================
# LISTADO DEL KARDEX CON FILTROS
# =======================================
def kardex_list(request):
    productos = Producto.objects.all().order_by("nombre")
    movimientos = MovimientoInventario.objects.select_related('producto').order_by('-fecha')

    producto_id = request.GET.get('producto')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Filtro por producto
    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)

    # Filtro por fechas — ajustado para DateTimeField
    if fecha_inicio:
        fi = parse_date(fecha_inicio)
        if fi:
            movimientos = movimientos.filter(fecha__gte=datetime.combine(fi, datetime.min.time()))
    if fecha_fin:
        ff = parse_date(fecha_fin)
        if ff:
            movimientos = movimientos.filter(fecha__lt=datetime.combine(ff + timedelta(days=1), datetime.min.time()))

    context = {
        'productos': productos,
        'movimientos': movimientos,
        'producto_seleccionado': producto_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    return render(request, 'kardex/kardex_list.html', context)



# =======================================
#  EXPORTAR KARDEX A EXCEL
# =======================================
def kardex_export_excel(request):
    # Mismos filtros que en kardex_list
    movimientos = MovimientoInventario.objects.select_related('producto').order_by('-fecha')

    producto_id = request.GET.get('producto')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)

    if fecha_inicio:
        fi = parse_date(fecha_inicio)
        if fi:
            movimientos = movimientos.filter(fecha__gte=datetime.combine(fi, datetime.min.time()))
    if fecha_fin:
        ff = parse_date(fecha_fin)
        if ff:
            movimientos = movimientos.filter(fecha__lt=datetime.combine(ff + timedelta(days=1), datetime.min.time()))

    # =======================
    #  CREAR ARCHIVO EXCEL
    # =======================
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = "KARDEX DE PRODUCTOS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Filtros aplicados
    filtro_texto = []
    if producto_id:
        prod = Producto.objects.filter(pk=producto_id).first()
        if prod:
            filtro_texto.append(f"Producto: {prod.nombre}")
    if fecha_inicio:
        filtro_texto.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filtro_texto.append(f"Hasta: {fecha_fin}")
    ws.merge_cells("A2:F2")
    ws["A2"] = " | ".join(filtro_texto) if filtro_texto else "Sin filtros"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = ["Fecha", "Tipo", "Cantidad", "Saldo", "Referencia", "Usuario"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Datos
    row = 4
    for m in movimientos:
        fecha_valor = m.fecha
        #  Convertir a naive datetime si tiene zona horaria
        if is_aware(fecha_valor):
            fecha_valor = make_naive(fecha_valor)

        ws.cell(row=row, column=1, value=fecha_valor)
        ws.cell(row=row, column=2, value=m.tipo)
        ws.cell(row=row, column=3, value=float(m.cantidad))
        ws.cell(row=row, column=4, value=float(getattr(m, "saldo", 0)))
        ws.cell(row=row, column=5, value=str(getattr(m, "referencia", "")))
        ws.cell(row=row, column=6, value=str(getattr(m, "usuario", "")))
        row += 1

    # Formatos
    for r in range(4, row):
        ws.cell(row=r, column=1).number_format = "dd/mm/yyyy hh:mm"
        ws.cell(row=r, column=3).number_format = numbers.FORMAT_NUMBER_00
        ws.cell(row=r, column=4).number_format = numbers.FORMAT_NUMBER_00

    # Ajuste de ancho
    widths = [20, 12, 12, 12, 40, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Descargar
    nombre = "kardex.xlsx"
    if producto_id:
        nombre = f"kardex_producto_{producto_id}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response
